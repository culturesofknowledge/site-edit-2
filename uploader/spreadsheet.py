import logging
from typing import Union, List, Set, Type

from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from uploader.constants import MANDATORY_SHEETS
from uploader.entities.entity import CofkEntity
from uploader.entities.locations import CofkLocations, CofkBulkLocations
from uploader.entities.manifestations import CofkManifestations
from uploader.entities.people import CofkPeople, CofkBulkPeople
from uploader.entities.repositories import CofkRepositories
from uploader.entities.work import CofkWork
from uploader.models import CofkCollectUpload

log = logging.getLogger(__name__)


class CofkExcelFileError(Exception):
    msg: str

    def __init__(self, msg: str):
        super().__init__(msg)
        self.msg = msg


class CofkSheet:
    def __init__(self, sheet: Worksheet):
        self.worksheet: Worksheet = sheet
        self.header: List[str]
        self.rows: int
        self.name: str = sheet.title
        self.entities: Union[Type[CofkEntity], None] = None
        # Hard-coded number of header rows
        self.header_length: int = 2

        # Obtain header and row count of non-empty rows
        rows = (row for row in self.worksheet.iter_rows() if any([cell.value is not None for cell in row]))
        self.header = [cell.value for cell in next(rows) if cell.value is not None]
        next(rows)
        self.rows = sum(1 for _ in rows)

    @property
    def missing_columns(self) -> Set[str]:
        return set(MANDATORY_SHEETS[self.name]['columns']).difference(set(self.header))


class CofkUploadExcelFile:

    def __init__(self, upload: CofkCollectUpload, filename: str):
        """
        :param filename:
        """
        self.errors = {}
        self.wb: Workbook | None = None
        self.works = None
        self.upload = upload
        self.filename = filename
        self.total_errors = 0
        self.sheets: dict[str: CofkSheet] = {}
        self.missing: List[str] = []

        self.wb = load_workbook(filename=filename, data_only=True, read_only=True)

        self.upload_type = self.check_sheets()

        if self.upload_type == 'work':
            self._process_work_upload()
        elif self.upload_type == 'people':
            self._process_people_only()
        elif self.upload_type == 'location':
            self._process_locations_only()

    def check_sheets(self) -> str:
        """Determine upload type from sheets present. Returns 'work', 'people', or 'location'."""
        present = set(self.wb.sheetnames)

        if 'Work' in present or 'Manifestation' in present:
            missing = set(MANDATORY_SHEETS.keys()) - present
            if missing:
                msg = f'Missing sheets: {", ".join(sorted(missing))}'
                log.error(msg)
                raise CofkExcelFileError(msg)
            return 'work'
        elif 'People' in present and 'Places' not in present:
            return 'people'
        elif 'Places' in present and 'People' not in present:
            return 'location'
        else:
            msg = ('Could not determine upload type. File must contain either a Work sheet '
                   '(standard upload), only a People sheet (bulk people), or only a Places sheet '
                   '(bulk locations).')
            log.error(msg)
            raise CofkExcelFileError(msg)

    def _load_sheet(self, sheet_name: str):
        """Load and validate a single sheet, appending to self.missing on error."""
        try:
            self.sheets[sheet_name] = CofkSheet(self.wb[sheet_name])
        except StopIteration:
            self.missing.append(f'{sheet_name} is missing its header rows.')
            return

        if self.sheets[sheet_name].missing_columns:
            cols = self.sheets[sheet_name].missing_columns
            ms = ('columns ' + ', '.join(cols)) if len(cols) > 1 else f'column "{cols.pop()}"'
            self.missing.append(f'Missing {ms} from the sheet {sheet_name}.')

    def _process_work_upload(self):
        """Process a standard 5-sheet work upload."""
        for sheet in MANDATORY_SHEETS.keys():
            self._load_sheet(sheet)

        if self.missing:
            raise CofkExcelFileError('</br> '.join(self.missing))

        if self.sheets['Work'].rows == 0:
            raise CofkExcelFileError("Spreadsheet contains no works.")

        sheets = ', '.join([
            f'{s}: {self.sheets[s].rows} ({self.sheets[s].worksheet.calculate_dimension()})'
            for s in MANDATORY_SHEETS.keys()
        ])
        log.info(f'{self.upload}: all {len(MANDATORY_SHEETS)} sheets verified: [{sheets}]')

        self.sheets['Repositories'].entities = CofkRepositories(upload=self.upload,
                                                                sheet=self.sheets['Repositories'])
        self.sheets['Places'].entities = CofkLocations(upload=self.upload, sheet=self.sheets['Places'],
                                                       work_data=self.sheets['Work'].worksheet)
        self.sheets['People'].entities = CofkPeople(upload=self.upload, sheet=self.sheets['People'])
        self.sheets['Work'].entities = CofkWork(upload=self.upload, sheet=self.sheets['Work'],
                                                people=self.sheets['People'].entities.people,
                                                locations=self.sheets['Places'].entities.locations)
        self.sheets['Manifestation'].entities = CofkManifestations(
            upload=self.upload,
            sheet=self.sheets['Manifestation'],
            repositories=self.sheets['Repositories'].entities.institutions,
            works=self.sheets['Work'].entities.works)

        if self.sheets['People'].entities.other_errors:
            for row_index in self.sheets['People'].entities.other_errors:
                for error in self.sheets['People'].entities.other_errors[row_index]:
                    if error['entity'] == 'work':
                        self.sheets['Work'].entities.add_error(error['error'], None, row_index)

        if self.sheets['Work'].entities.errors:
            self.errors['work'] = self.sheets['Work'].entities.format_errors_for_template()
            self.total_errors += self.errors['work']['total']
        if self.sheets['People'].entities.errors:
            self.errors['people'] = self.sheets['People'].entities.format_errors_for_template()
            self.total_errors += self.errors['people']['total']
        if self.sheets['Repositories'].entities.errors:
            self.errors['repositories'] = self.sheets['Repositories'].entities.format_errors_for_template()
            self.total_errors += self.errors['repositories']['total']
        if self.sheets['Places'].entities.errors:
            self.errors['locations'] = self.sheets['Places'].entities.format_errors_for_template()
            self.total_errors += self.errors['locations']['total']
        if self.sheets['Manifestation'].entities and self.sheets['Manifestation'].entities.errors:
            self.errors['manifestations'] = self.sheets['Manifestation'].entities.format_errors_for_template()
            self.total_errors += self.errors['manifestations']['total']

        if self.total_errors == 0:
            people = self.sheets['People'].entities.people
            self.sheets['People'].entities.bulk_create(people)
            self.sheets['People'].entities.log_summary.append(f'{len(people)} CofkCollectPerson')

            locations = self.sheets['Places'].entities.locations
            self.sheets['Places'].entities.bulk_create(locations)
            self.sheets['Places'].entities.log_summary.append(f'{len(locations)} CofkCollectLocation')

            self.sheets['Work'].entities.create_all()

            inst = self.sheets['Repositories'].entities.institutions
            self.sheets['Repositories'].entities.bulk_create(inst)
            self.sheets['Repositories'].entities.log_summary.append(f'{len(inst)} CofkCollectInstitution')

            manifs = self.sheets['Manifestation'].entities.manifestations
            self.sheets['Manifestation'].entities.bulk_create(manifs)
            self.sheets['Manifestation'].entities.log_summary.append(f'{len(manifs)} CofkCollectManifestation')

            log_msg = (self.sheets['Work'].entities.log_summary
                       + self.sheets['Manifestation'].entities.log_summary
                       + self.sheets['Repositories'].entities.log_summary)
            log.info(f'{self.upload}: created ' + ', '.join(log_msg))

    def _process_people_only(self):
        """Process a people-only upload.

        Detects format automatically:
        - Standard (3-column) format: header row contains 'primary_name'
        - Bulk (24-column) format: header row contains verbose column titles
        """
        try:
            sheet = CofkSheet(self.wb['People'])
        except StopIteration:
            raise CofkExcelFileError('People sheet is missing its header rows.')

        is_bulk = 'primary_name' not in sheet.header

        if not is_bulk and sheet.missing_columns:
            cols = sheet.missing_columns
            ms = ('columns ' + ', '.join(cols)) if len(cols) > 1 else f'column "{cols.pop()}"'
            raise CofkExcelFileError(f'Missing {ms} from the People sheet.')

        self.sheets['People'] = sheet
        entity_class = CofkBulkPeople if is_bulk else CofkPeople
        self.sheets['People'].entities = entity_class(upload=self.upload, sheet=self.sheets['People'])

        if self.sheets['People'].entities.errors:
            self.errors['people'] = self.sheets['People'].entities.format_errors_for_template()
            self.total_errors += self.errors['people']['total']

        if self.total_errors == 0:
            people = self.sheets['People'].entities.people
            self.sheets['People'].entities.bulk_create(people)
            fmt = 'bulk people' if is_bulk else 'people'
            log.info(f'{self.upload}: created {len(people)} CofkCollectPerson ({fmt} upload)')

    def _process_locations_only(self):
        """Process a locations-only upload.

        Detects format automatically:
        - Standard (2-column) format: header row contains 'location_name'
        - Bulk (14-column) format: header row contains verbose column titles
        """
        try:
            sheet = CofkSheet(self.wb['Places'])
        except StopIteration:
            raise CofkExcelFileError('Places sheet is missing its header rows.')

        is_bulk = 'location_name' not in sheet.header

        if not is_bulk and sheet.missing_columns:
            cols = sheet.missing_columns
            ms = ('columns ' + ', '.join(cols)) if len(cols) > 1 else f'column "{cols.pop()}"'
            raise CofkExcelFileError(f'Missing {ms} from the Places sheet.')

        self.sheets['Places'] = sheet

        if is_bulk:
            self.sheets['Places'].entities = CofkBulkLocations(upload=self.upload,
                                                               sheet=self.sheets['Places'])
        else:
            self.sheets['Places'].entities = CofkLocations(upload=self.upload,
                                                           sheet=self.sheets['Places'],
                                                           work_data=None)

        if self.sheets['Places'].entities.errors:
            self.errors['locations'] = self.sheets['Places'].entities.format_errors_for_template()
            self.total_errors += self.errors['locations']['total']

        if self.total_errors == 0:
            locations = self.sheets['Places'].entities.locations
            self.sheets['Places'].entities.bulk_create(locations)
            fmt = 'bulk locations' if is_bulk else 'locations'
            log.info(f'{self.upload}: created {len(locations)} CofkCollectLocation ({fmt} upload)')
