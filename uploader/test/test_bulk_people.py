import logging
import tempfile

from openpyxl.workbook import Workbook

from uploader.models import CofkCollectPerson
from uploader.spreadsheet import CofkUploadExcelFile
from uploader.test.test_serv import UploadIncludedTestCase
from uploader.validation import CofkExcelFileError

log = logging.getLogger(__name__)

# Verbose column headers that trigger bulk detection (no 'primary_name' in row 1)
BULK_PEOPLE_HEADERS = [
    'Primary name', 'Synonyms', 'Occupations/roles/titles', 'GENDER', 'IS ORGANIZATION',
    'BIRTH YEAR', 'BIRTH YEAR INFERRED', 'BIRTH YEAR UNCERTAIN', 'BIRTH YEAR APPROX',
    'DEATH YEAR', 'DEATH YEAR INFERRED', 'DEATH YEAR UNCERTAIN', 'DEATH YEAR APPROX',
    'FLOURISHED EARLIEST YEAR 1', 'FLOURISHED LATEST YEAR 2', 'FLOURISHED IS RANGE',
    'FLOURISHED YEAR INFERRED', 'FLOURISHED YEAR UNCERTAIN', 'FLOURISHED YEAR APPROX',
    'GENERAL NOTES ON PERSON', "EDITORS' NOTES AND QUERIES",
    'RELATED RESOURCE NAME', 'RELATED RESOURCE URL', 'Further reading',
]


class TestBulkPeople(UploadIncludedTestCase):

    def create_bulk_people_file(self, data_rows: list) -> str:
        wb = Workbook()
        ws = wb.create_sheet('People')
        ws.append(BULK_PEOPLE_HEADERS)
        ws.append(['-'] * len(BULK_PEOPLE_HEADERS))
        for row in data_rows:
            ws.append(row)
        tf = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        wb.save(tf.name)
        self.tmp_files.append(tf.name)
        return tf.name

    def _row(self, primary_name=None, alternative_names=None, roles=None, gender=None,
             is_org=None, birth_year=None, birth_inferred=None, birth_uncertain=None,
             birth_approx=None, death_year=None, death_inferred=None, death_uncertain=None,
             death_approx=None, fl_year=None, fl2_year=None, fl_range=None,
             fl_inferred=None, fl_uncertain=None, fl_approx=None,
             notes=None, editors_notes=None, resource_name=None, resource_url=None,
             further_reading=None):
        return [primary_name, alternative_names, roles, gender, is_org,
                birth_year, birth_inferred, birth_uncertain, birth_approx,
                death_year, death_inferred, death_uncertain, death_approx,
                fl_year, fl2_year, fl_range, fl_inferred, fl_uncertain, fl_approx,
                notes, editors_notes, resource_name, resource_url, further_reading]

    def test_bulk_people_creates_records(self):
        """Valid bulk people upload creates CofkCollectPerson records with correct field values."""
        row = self._row(
            primary_name='Isaac Newton',
            gender='M',
            birth_year=1643,
            death_year=1727,
            editors_notes='test notes',
        )
        filename = self.create_bulk_people_file([row])

        cuef = CofkUploadExcelFile(self.new_upload, filename)

        self.assertEqual(cuef.errors, {})
        self.assertEqual(CofkCollectPerson.objects.count(), 1)
        person = CofkCollectPerson.objects.first()
        self.assertEqual(person.primary_name, 'Isaac Newton')
        self.assertEqual(person.gender, 'M')
        self.assertEqual(person.date_of_birth_year, 1643)
        self.assertEqual(person.date_of_death_year, 1727)
        self.assertEqual(person.editors_notes, 'test notes')

    def test_bulk_people_multiple_records(self):
        """Multiple rows create multiple CofkCollectPerson records."""
        rows = [
            self._row(primary_name='Isaac Newton', birth_year=1643),
            self._row(primary_name='Robert Boyle', birth_year=1627),
            self._row(primary_name='John Locke', death_year=1704),
        ]
        filename = self.create_bulk_people_file(rows)

        cuef = CofkUploadExcelFile(self.new_upload, filename)

        self.assertEqual(cuef.errors, {})
        self.assertEqual(CofkCollectPerson.objects.count(), 3)

    def test_bulk_people_minimal(self):
        """Only primary_name is required; upload succeeds with all other fields omitted."""
        filename = self.create_bulk_people_file([self._row(primary_name='John Doe')])

        cuef = CofkUploadExcelFile(self.new_upload, filename)

        self.assertEqual(cuef.errors, {})
        self.assertEqual(CofkCollectPerson.objects.count(), 1)

    def test_bulk_people_missing_required(self):
        """Row without primary_name generates a validation error and no record is created."""
        row = self._row(birth_year=1600)
        filename = self.create_bulk_people_file([row])

        cuef = CofkUploadExcelFile(self.new_upload, filename)

        self.assertIn('people', cuef.errors)
        self.assertEqual(CofkCollectPerson.objects.count(), 0)

    def test_bulk_people_duplicate_skipped(self):
        """Two rows with the same name (case-insensitive) result in only one record."""
        rows = [
            self._row(primary_name='Isaac Newton'),
            self._row(primary_name='isaac newton'),
        ]
        filename = self.create_bulk_people_file(rows)

        cuef = CofkUploadExcelFile(self.new_upload, filename)

        self.assertEqual(cuef.errors, {})
        self.assertEqual(CofkCollectPerson.objects.count(), 1)

    def test_bulk_people_invalid_bool(self):
        """A non-0/1 value for a boolean field generates a validation error."""
        row = self._row(primary_name='Isaac Newton', birth_inferred=2)
        filename = self.create_bulk_people_file([row])

        cuef = CofkUploadExcelFile(self.new_upload, filename)

        self.assertIn('people', cuef.errors)
        self.assertEqual(CofkCollectPerson.objects.count(), 0)

    def test_bulk_people_year_below_minimum(self):
        """A year below MIN_YEAR (1400) generates a validation error."""
        row = self._row(primary_name='Isaac Newton', birth_year=1200)
        filename = self.create_bulk_people_file([row])

        cuef = CofkUploadExcelFile(self.new_upload, filename)

        self.assertIn('people', cuef.errors)

    def test_bulk_people_year_above_maximum(self):
        """A year above MAX_YEAR (1900) generates a validation error."""
        row = self._row(primary_name='Someone Recent', death_year=1950)
        filename = self.create_bulk_people_file([row])

        cuef = CofkUploadExcelFile(self.new_upload, filename)

        self.assertIn('people', cuef.errors)

    def test_bulk_people_bool_fields_accepted(self):
        """Valid boolean values 0 and 1 are accepted and stored correctly."""
        row = self._row(
            primary_name='Isaac Newton',
            birth_inferred=1,
            birth_uncertain=0,
            birth_approx=1,
            death_inferred=0,
            fl_range=1,
        )
        filename = self.create_bulk_people_file([row])

        cuef = CofkUploadExcelFile(self.new_upload, filename)

        self.assertEqual(cuef.errors, {})
        person = CofkCollectPerson.objects.first()
        self.assertEqual(person.date_of_birth_inferred, 1)
        self.assertEqual(person.date_of_birth_uncertain, 0)
        self.assertEqual(person.date_of_birth_approx, 1)
        self.assertEqual(person.flourished_is_range, 1)

    def test_bulk_people_flourished_dates(self):
        """Flourished year range is stored correctly."""
        row = self._row(primary_name='Anonymous Poet', fl_year=1550, fl2_year=1600, fl_range=1)
        filename = self.create_bulk_people_file([row])

        cuef = CofkUploadExcelFile(self.new_upload, filename)

        self.assertEqual(cuef.errors, {})
        person = CofkCollectPerson.objects.first()
        self.assertEqual(person.flourished_year, 1550)
        self.assertEqual(person.flourished2_year, 1600)
        self.assertEqual(person.flourished_is_range, 1)

    def test_bulk_people_upload_type_detected(self):
        """A People-only bulk spreadsheet is detected as upload_type 'people'."""
        filename = self.create_bulk_people_file([self._row(primary_name='Isaac Newton')])

        cuef = CofkUploadExcelFile(self.new_upload, filename)

        self.assertEqual(cuef.upload_type, 'people')

    def test_both_people_and_places_without_work_raises_error(self):
        """A file with both People and Places sheets but no Work sheet raises CofkExcelFileError."""
        wb = Workbook()
        wb.create_sheet('People')
        wb.create_sheet('Places')
        tf = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        wb.save(tf.name)
        self.tmp_files.append(tf.name)

        with self.assertRaises(CofkExcelFileError):
            CofkUploadExcelFile(self.new_upload, tf.name)
