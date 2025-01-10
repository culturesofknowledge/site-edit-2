import logging
import tempfile
import zipfile

from openpyxl.workbook import Workbook

from institution.models import CofkUnionInstitution
from location.models import CofkUnionLocation
from person.models import CofkUnionPerson
from uploader.constants import MANDATORY_SHEETS
from uploader.models import CofkCollectWork, CofkCollectAuthorOfWork, \
    CofkCollectAddresseeOfWork, CofkCollectOriginOfWork, CofkCollectDestinationOfWork, CofkCollectManifestation
from uploader.spreadsheet import CofkUploadExcelFile
from uploader.test.test_serv import UploadIncludedTestCase, spreadsheet_data
from uploader.validation import CofkExcelFileError

log = logging.getLogger(__name__)


class TestFileUpload(UploadIncludedTestCase):
    def test_create_upload(self):
        self.assertEqual(self.new_upload.upload_status.status_desc, 'Awaiting review')
        self.assertEqual(self.new_upload.total_works, 0)
        self.assertEqual(self.new_upload.works_accepted, 0)
        self.assertEqual(self.new_upload.works_rejected, 0)

    def test_non_excel_file(self):
        tf = tempfile.NamedTemporaryFile(suffix='.xlsx')
        msg = 'File is not a zip file'
        self.assertRaisesRegex(zipfile.BadZipFile, msg, CofkUploadExcelFile,
                               self.new_upload, tf.name)

    def test_empty_file(self):
        wb = Workbook()
        tf = tempfile.NamedTemporaryFile(suffix='.xlsx')
        wb.save(tf.name)

        msg = r'Missing sheets: Manifestation, People, Places, Repositories, Work'
        self.assertRaisesRegex(CofkExcelFileError, msg, CofkUploadExcelFile,
                               self.new_upload, tf.name)

    def test_incomplete_headers(self):
        wb = Workbook()

        for sheet_name in MANDATORY_SHEETS.keys():
            ws = wb.create_sheet(sheet_name)
            ws.append(MANDATORY_SHEETS[sheet_name]['columns'])

        tf = tempfile.NamedTemporaryFile(suffix='.xlsx')
        wb.save(tf.name)

        msg = r'People is missing its header rows.'
        self.assertRaisesRegex(CofkExcelFileError, msg, CofkUploadExcelFile,
                               self.new_upload, tf.name)

    def test_no_data(self):
        filename = self.create_excel_file()

        msg = 'Spreadsheet contains no works.'
        self.assertRaisesRegex(CofkExcelFileError, msg, CofkUploadExcelFile,
                               self.new_upload, filename)

    def test_work_data(self):
        work_data = [
            [1, "test", "J", 1660, 1, 1, 1660, 1, 2, 1, 1, 1, 1, "test", "newton", 15257, "test", 1, 1, "test", "Wren",
             22859, "test", 1, 1, "test", "Burford", 400285, "test", 1, 1, "Carisbrooke", 782, "test", 1, 1, "test",
             "", "fra;eng", '', '', '', '', '', '', "test", "test", "test", "Baskerville", 885, "test",
             "test", "EMLO", "http://emlo.bodleian.ox.ac.uk/", "Early Modern Letters Online test"]]
        filename = self.create_excel_file({'Work': work_data})

        msg = 'Person with the id 15257 was listed in the Work sheet but is not present in the People sheet.'
        msg_2 = 'Location with the id 400285 was listed in the Work sheet but is not present in the Places sheet.'
        msg_3 = 'Person with the id 885 was listed in the Work sheet but is not present in the People sheet.'
        msg_4 = 'Person with the id 22859 was listed in the Work sheet but is not present in the People sheet.'
        msg_5 = 'Location with the id 782 was listed in the Work sheet but is not present in the Places sheet.'

        cuef = CofkUploadExcelFile(self.new_upload, filename)

        self.assertEqual(cuef.errors['work']['total'], 5)
        self.assertIn(msg, cuef.errors['work']['errors'][0]['errors'])
        self.assertIn(msg_2, cuef.errors['work']['errors'][0]['errors'])
        self.assertIn(msg_3, cuef.errors['work']['errors'][0]['errors'])
        self.assertIn(msg_4, cuef.errors['work']['errors'][0]['errors'])
        self.assertIn(msg_5, cuef.errors['work']['errors'][0]['errors'])

    def test_successful_data(self):
        """
        This test should run successfully as all required data is present and valid.
        """
        filename = self.create_excel_file(spreadsheet_data)

        cuef = CofkUploadExcelFile(self.new_upload, filename)

        # This is a valid upload and should be without errors
        self.assertEqual(cuef.errors, {})
        self.assertEqual(CofkCollectWork.objects.count(), 1)
        self.assertEqual(CofkCollectAuthorOfWork.objects.count(), 1)
        self.assertEqual(CofkCollectAddresseeOfWork.objects.count(), 1)
        self.assertEqual(CofkCollectOriginOfWork.objects.count(), 1)
        self.assertEqual(CofkCollectDestinationOfWork.objects.count(), 1)
        self.assertEqual(CofkCollectManifestation.objects.count(), 2)

    def test_nonsense(self):
        """
        This test tries to import a work with a non-ISO639 language "aaj", and other invalid data
        """
        data = {'Work': [[1, "test", "J", 1660, 1, 1, 'sss', 1, 2, 1, 1, 1, 1, "test", "newton", 15257, "test", 1, 1,
                          "test", "Wren", 22859, "test", 1, "s1", "test", "Burford", 1, "test", 1, 1, "Carisbrooke",
                          782, "test", 1, 1, "test", "", "fra;aaj", '', '', '', '', '', '', "test", "test", "test",
                          "Baskerville", 885, "test", "test", "EMLO", "http://emlo.bodleian.ox.ac.uk/",
                          "Early Modern Letters Online test"]],
                'Places': [['Burford', 1],
                           ['Carisbrooke', 782]],
                'Manifestation': [[1, 1, "ALS", 2, "Bodleian", "test", "test", '', '', '', '', ''],
                                  [2, 1, '', '', '', '', '', "P", "test", "test", '', '']],
                'Repositories': [['Bodleian', 2]]
                }

        filename = self.create_excel_file(data)

        msg = 'The value in column "language_id", "aaj" is not a valid ISO639 language.'
        msg_2 = 'Column date_of_work2_std_year in Work sheet is not a valid integer (value: sss).'
        msg_3 = 'Column addressees_uncertain in Work sheet is not a boolean value of either 0 or 1 (value: s1).'
        msg_4 = 'There is no location with the id 1 in the Union catalogue.'
        msg_5 = 'There is no repository with the id 2 in the Union catalogue.'
        cuef = CofkUploadExcelFile(self.new_upload, filename)

        self.assertEqual(cuef.errors['work']['total'], 7)
        self.assertIn(msg, cuef.errors['work']['errors'][0]['errors'])
        self.assertIn(msg_2, cuef.errors['work']['errors'][0]['errors'])
        self.assertIn(msg_3, cuef.errors['work']['errors'][0]['errors'])
        self.assertIn(msg_4, cuef.errors['locations']['errors'][0]['errors'])
        self.assertIn(msg_5, cuef.errors['manifestations']['errors'][0]['errors'])

    def test_mismatching_people(self):
        """
        This test tries to import a work and a people sheet that has one name,
        but two ids neither of which exists in the Union catalogue. The test insures
        that the data is properly parsed and processed from the spreadsheet.
        """
        data = {'Work': [[1, "test", "J", 1660, 1, 1, 'sss', 1, 2, 1, 1, 1, 1, "test", "newton", 15257, "test", 1, 1,
                          "test", "Wren", 22859, "test", 1, "s1", "test", "Burford", 1, "test", 1, 1, "Carisbrooke",
                          782, "test", 1, 1, "test", "", "fra;aaj", '', '', '', '', '', '', "test", "test", "test",
                          "Baskerville", 885, "test", "test", "EMLO", "http://emlo.bodleian.ox.ac.uk/",
                          "Early Modern Letters Online test"]],
                'Places': [['Burford', 1],
                           ['Carisbrooke', 782]],
                'People': [["Baskerville", 885],
                           ["newton", 15257],
                           ["Wren", 22859],
                           ["I. et J. Beeckman", "903506;908149"]],
                'Manifestation': [[1, 1, "ALS", 2, "Bodleian", "test", "test", '', '', '', '', ''],
                                  [2, 1, '', '', '', '', '', "P", "test", "test", '', '']],
                'Repositories': [['Bodleian', 2]]
                }

        filename = self.create_excel_file(data)

        cuef = CofkUploadExcelFile(self.new_upload, filename)

        msg = 'There is no person with the id 903506 in the Union catalogue.'
        msg2 = 'There is no person with the id 908149 in the Union catalogue.'

        self.assertIn(msg, cuef.errors['people']['errors'][0]['errors'])
        self.assertIn(msg2, cuef.errors['people']['errors'][0]['errors'])


    def test_mismatching_people_one_exists(self):
        """
        This test replicates the above test except that one of the ids exists in the Union catalogue.
        """
        CofkUnionPerson.objects.create(iperson_id=903506, foaf_name='I. et J. Beeckman')

        data = {'Work': [[1, "test", "J", 1660, 1, 1, 'sss', 1, 2, 1, 1, 1, 1, "test", "newton", 15257, "test", 1, 1,
                          "test", "Wren", 22859, "test", 1, "s1", "test", "Burford", 1, "test", 1, 1, "Carisbrooke",
                          782, "test", 1, 1, "test", "", "fra;aaj", '', '', '', '', '', '', "test", "test", "test",
                          "Baskerville", 885, "test", "test", "EMLO", "http://emlo.bodleian.ox.ac.uk/",
                          "Early Modern Letters Online test"]],
                'Places': [['Burford', 1],
                           ['Carisbrooke', 782]],
                'People': [["Baskerville", 885],
                           ["newton", 15257],
                           ["Wren", 22859],
                           ["I. et J. Beeckman", "903506;908149"]],
                'Manifestation': [[1, 1, "ALS", 2, "Bodleian", "test", "test", '', '', '', '', ''],
                                  [2, 1, '', '', '', '', '', "P", "test", "test", '', '']],
                'Repositories': [['Bodleian', 2]]
                }

        filename = self.create_excel_file(data)

        cuef = CofkUploadExcelFile(self.new_upload, filename)

        msg = 'There is no person with the id 908149 in the Union catalogue.'
        msg2 = 'There is no person with the id 903506 in the Union catalogue.'

        self.assertIn(msg, cuef.errors['people']['errors'][0]['errors'])
        self.assertNotIn(msg2, cuef.errors['work']['errors'][0]['errors'])

    def test_location_created(self):
        """
        This test is similar to the above two tests except that it tests the creation of two locations.
        """
        CofkUnionLocation.objects.create(location_id=1, location_name='Burford')
        CofkUnionInstitution.objects.create(institution_id=2, institution_name='Bodleian')

        data = {
            'Work': [[1, "test", "J", 1660, 1, 1, 'sss', 1, 2, 1, 1, 1, 1, "test", "newton", 15257, "test", 1, 1,
                      "test", "Wren", 22859, "test", 1, "s1", "test", "Burford", 1, "test", 1, 1, "Carisbrooke",
                      782, "test", 1, 1, "test", "", "fra;aaj", '', '', '', '', '', '', "test", "test", "test",
                      "Baskerville", 885, "test", "test", "EMLO", "http://emlo.bodleian.ox.ac.uk/",
                      "Early Modern Letters Online test"]],
            'Places': [['Burford', 1],
                       ['Carisbrooke', 782]],
            'Manifestation': [[1, 1, "ALS", 2, "Bodleian", "test", "test", '', '', '', '', ''],
                              [2, 1, '', '', '', '', '', "P", "test", "test", '', '']],
            'Repositories': [['Bodleian', 2]]
            }

        filename = self.create_excel_file(data)

        msg = 'There is no location with the id 1 in the Union catalogue.'
        msg_2 = 'There is no repository with the id 2 in the Union catalogue.'
        cuef = CofkUploadExcelFile(self.new_upload, filename)

        self.assertNotIn(msg, cuef.errors['work']['errors'][0]['errors'])
        self.assertNotIn(msg_2, cuef.errors['work']['errors'][0]['errors'])