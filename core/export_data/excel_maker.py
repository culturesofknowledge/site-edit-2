"""
This module contains the functions to create Excel file and object(Workbook)
for modules (e.g. work, person, location, etc.)
"""
import itertools
import logging
from typing import Iterable, NoReturn, Callable

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from core import constant
from core.export_data import excel_header_values, excel_serv
from core.helper import model_serv
from core.helper.view_components import HeaderValues
from core.models import CofkUnionResource
from sharedlib import data_utils
from work.models import CofkUnionWork

log = logging.getLogger(__name__)


def fill_sheet(sheet: 'Worksheet',
               rows: Iterable[CofkUnionWork],
               header_values: HeaderValues,
               sheet_name) -> NoReturn:
    sheet.title = sheet_name

    # fill header
    sheet.append(header_values.get_header_list())

    apply_header_style(sheet)
    excel_serv.fill_header_style(sheet)

    rows = (header_values.obj_to_values(obj) for obj in rows)
    rows = map(data_utils.to_str_list_no_none, rows)
    rows = map(excel_serv.escape_xlsx_char_by_row, rows)
    for row in rows:
        sheet.append(row)


def fill_work_sheet(sheet, rows):
    log.debug('start fill work sheet')
    return fill_sheet(sheet, rows=rows,
                      header_values=excel_header_values.WorkExcelHeaderValues(),
                      sheet_name='work')


def fill_person_sheet(sheet, rows):
    log.debug('start fill person sheet')
    return fill_sheet(sheet, rows=rows,
                      header_values=excel_header_values.PersonExcelHeaderValues(),
                      sheet_name='person')


def fill_location_sheet(sheet, rows):
    log.debug('start fill location sheet')
    return fill_sheet(sheet, rows=rows,
                      header_values=excel_header_values.LocationExcelHeaderValues(),
                      sheet_name='location')


def fill_manif_sheet(sheet, rows):
    log.debug('start fill manifestation sheet')
    return fill_sheet(sheet, rows=rows,
                      header_values=excel_header_values.ManifExcelHeaderValues(),
                      sheet_name='manifestation')


def fill_inst_sheet(sheet, rows):
    log.debug('start fill institution sheet')
    return fill_sheet(sheet, rows=rows,
                      header_values=excel_header_values.InstExcelHeaderValues(),
                      sheet_name='institution')


def fill_resource_sheet(sheet, rows):
    log.debug('start fill resource sheet')
    return fill_sheet(sheet, rows=rows,
                      header_values=excel_header_values.ResourceExcelHeaderValues(),
                      sheet_name='resource')


def apply_header_style(sheet):
    # apply style font bold and background color gray in the first row
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type='solid', fgColor='D3D3D3')


def get_flat_resource_list(objects, get_resource_map_set_fn) -> Iterable['CofkUnionResource']:
    return itertools.chain.from_iterable((r.resource for r in get_resource_map_set_fn(w).iterator())
                                         for w in objects)


def _create_excel_by_fill_fn(fill_fn: Callable[[Workbook], NoReturn],
                             file_path: str = None):
    log.info('[Start] create excel')
    wb = openpyxl.Workbook()
    wb.remove_sheet(wb.active)

    fill_fn(wb)

    if file_path:
        wb.save(file_path)

    log.info(f'[Completed] create excel [{file_path=}] ')
    return wb


def create_work_excel(queryable_works: Iterable[CofkUnionWork],
                      file_path: str = None) -> 'openpyxl.Workbook':
    def _find_manif_list():
        manif_list = itertools.chain.from_iterable(w.manif_set.iterator()
                                                   for w in queryable_works)
        return model_serv.UniqueModelPkFilter(manif_list)

    def _find_person_list():
        person_list = itertools.chain.from_iterable(w.find_persons_by_rel_type([
            constant.REL_TYPE_CREATED,
            constant.REL_TYPE_SENT,
            constant.REL_TYPE_SIGNED,
            constant.REL_TYPE_WAS_ADDRESSED_TO,
            constant.REL_TYPE_INTENDED_FOR,
            constant.REL_TYPE_MENTION,
        ]) for w in queryable_works)
        return model_serv.UniqueModelPkFilter(person_list)

    def _find_location_list():
        location_list = itertools.chain.from_iterable(w.find_locations_by_rel_type([
            constant.REL_TYPE_WAS_SENT_FROM,
            constant.REL_TYPE_WAS_SENT_TO,
        ]) for w in queryable_works)
        return model_serv.UniqueModelPkFilter(location_list)

    def _find_inst_list():
        inst_list = (manif.inst for manif in model_serv.UniqueModelPkFilter(_find_manif_list()))
        inst_list = filter(None, inst_list)
        return model_serv.UniqueModelPkFilter(inst_list)

    def _fill_fn(workbook: Workbook):
        fill_work_sheet(workbook.create_sheet(), queryable_works)
        fill_person_sheet(workbook.create_sheet(), _find_person_list())
        fill_location_sheet(workbook.create_sheet(), _find_location_list())
        fill_manif_sheet(workbook.create_sheet(), _find_manif_list())
        fill_inst_sheet(workbook.create_sheet(), _find_inst_list())

        resource_list = itertools.chain(
            get_flat_resource_list(queryable_works, lambda obj: obj.cofkworkresourcemap_set),
            get_flat_resource_list(_find_person_list(), lambda obj: obj.cofkpersonresourcemap_set),
            get_flat_resource_list(_find_location_list(), lambda obj: obj.cofklocationresourcemap_set),
            get_flat_resource_list(_find_inst_list(), lambda obj: obj.cofkinstitutionresourcemap_set),
        )
        fill_resource_sheet(workbook.create_sheet(), resource_list)

    return _create_excel_by_fill_fn(_fill_fn, file_path=file_path)
